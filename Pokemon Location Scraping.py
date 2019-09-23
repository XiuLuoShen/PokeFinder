#!/usr/bin/env python
# coding: utf-8

# In[94]:


import requests
import csv
import re
import json
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border


# In[2]:


rootUrl = 'https://pokemondb.net'


# In[3]:


notAvailableText ='Trade/migrate from another game'


# In[48]:


def pokemonGamesList():
    coreGamesUrl = requests.get('https://bulbapedia.bulbagarden.net/wiki/Core_series')
    page = BeautifulSoup(coreGamesUrl.text, "html.parser")
    coreGamesUrl.close()
    return [a.text for a in page.select('table[class="roundy"] td [title*="Pok"]')][4:] # Skip the Japan only games


# In[49]:


games = pokemonGamesList()
gamesSet = set(games)


# In[6]:


def getAllPokemonUrls():
    dexUrl = rootUrl + '/pokedex/national'
    pokeUrlsList = requests.get(dexUrl)
    soup = BeautifulSoup(pokeUrlsList.text, "html.parser")
    pokeUrlsList.close()
    return [a.attrs.get( 'href' ) for a in soup.select( 'div [class="infocard"] a[class="ent-name"]' )]


# In[7]:


def getPokemonLocationData(pokemonUrl, pokemonDictionary):
    htmlData = requests.get(rootUrl + pokemonUrl)
    soup = BeautifulSoup(htmlData.text, "html.parser")
    htmlData.close()
    whereHeader = soup.find('h2', string=re.compile("Where to find"))

    gameData = {}
    gameDataTag = whereHeader.find_all_next('span', class_=re.compile("igame"))
    name = pokemonUrl.rsplit('/', 1)[-1] # A bit hacky
#     print(name) # Print the name of the pokemon. (for debugging)
    for tag in gameDataTag:
        if tag.text in gamesSet:
            if (tag.parent.parent.find('td')):
                # gameData[game] = location
                gameData[tag.text] = tag.parent.parent.find('td').text
    pokemonDictionary[name] = gameData


# In[8]:


def getAllPokemonInfo():
    """
    Scrape the pokemondb and get all pokemon location information
    Returns a dictionary
    """
    pokemonInfo = {}
    for pokeurl in getAllPokemonUrls():
        getPokemonLocationData(pokeurl, pokemonInfo)
    return pokemonInfo


# In[19]:


def saveToJSON():
    """
    Saves the data to a text file to avoid having to scrape for it again in the future.
    Fails if the file already exists.
    """
    try
        ofile = open('PokemonLocationData.json', 'x+')
        data = getAllPokemonInfo()
        json.dump(data, ofile)
    finally
        ofile.close()


# In[25]:


def loadFromJSON():
    """
    Load the data from a file that was previously saved.
    """
    # Open file with information that has already been scraped and saved as json data
    try:
        ifile = open('PokemonLocationData.json', 'r')
        pokemonData = json.loads(ifile.readline())
    finally:
        ifile.close()
    return pokemonData


# In[26]:


pokemonData = loadFromJSON()


# In[27]:


pokemonData['rayquaza'] # test


# In[107]:


def saveToExcel():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Pokemon Location in Games Data'

    redFill = PatternFill(start_color='ff6666', end_color='ff6666', fill_type='solid')
    greenFill = PatternFill(start_color='1aff8c', end_color='1aff8c', fill_type='solid')

    # Headers
    sheet['A1'] = '#'
    sheet['B1'] = 'Pokemon'
    for row in sheet.iter_rows(min_row=1, min_col=3, max_col=(len(games)+2), max_row=1):
        i = 0
        for cell in row:
            cell.value = games[i]
            cols.append(cell.column)
            i += 1

    # Fill in data
    number = 1
    for pokemon in list(pokemonData.keys()):
        for row in sheet.iter_rows(min_row=number+1, max_row=number+1, min_col=1, max_col=(len(games)+2)):
            row[0].value = str(number)
            row[1].value = pokemon.capitalize()
            i = 0
            for cell in row[2:]:
                # If the pokemon did not exist yet in that game or is unavailable, leave the cell as null and color it red
                if games[i] not in pokemonData[pokemon] or pokemonData[pokemon][games[i]] == notAvailableText:
                    cell.value = ''
                    cell.fill = redFill
                else:
                    cell.value = pokemonData[pokemon][games[i]]
                    cell.fill = greenFill
                i += 1
        number += 1

    wb.save('PokemonLocations.xlsx')


# In[110]:


# saveToExcel()


# In[ ]:




